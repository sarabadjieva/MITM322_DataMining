from pathlib import Path
from text_utils import normalize_text
import openpyxl
from territory_utils import canonical_area_name, is_total_row
import const_vals as cvals

def find_first_data_row(ws):
    """
    Find the first row where column A contains 'Общо за страната'
    or equivalent total label. This is treated as the first data row.
    """
    for r in range(1, ws.max_row + 1):
        name = normalize_text(ws.cell(r, 1).value)
        if is_total_row(name):
            return r
    return None

def find_rows_to_delete(ws, data_start):
    rows_to_delete = set()
    seen_oblasti = set()
    current_oblast = None

    for r in range(data_start, ws.max_row + 1):
        name = normalize_text(ws.cell(r, 1).value)
        if not name:
            continue

        if is_total_row(name):
            current_oblast = None
            continue

        canon = canonical_area_name(name)

        # first oblast row in a block -> keep
        if canon in cvals.AREA_LABELS and canon not in seen_oblasti:
            seen_oblasti.add(canon)
            current_oblast = canon
            continue

        # second identical oblast-name row inside block = municipality
        if current_oblast and canon == current_oblast:
            rows_to_delete.add(r)
            continue

        # any other non-oblast row is municipality/detail -> delete
        rows_to_delete.add(r)

    return rows_to_delete

def delete_rows_bottom_up(ws, rows_to_delete):
    for r in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(r, 1)


def clean_workbook_keep_only_oblasti(input_file, output_file=None, inplace=False) -> Path:
    input_file = Path(input_file)

    if inplace:
        output_file = input_file
    else:
        output_file = Path(output_file) if output_file else input_file.with_name(
            f"{input_file.stem}_oblasti_only{input_file.suffix}"
        )

    wb = openpyxl.load_workbook(input_file)
    summary = []

    for ws in wb.worksheets:
        data_start = find_first_data_row(ws)

        if data_start is None:
            summary.append((ws.title, "skipped_no_total_found", 0))
            continue

        rows_to_delete = find_rows_to_delete(ws, data_start)
        delete_rows_bottom_up(ws, rows_to_delete)

        summary.append((ws.title, "cleaned", len(rows_to_delete)))

    wb.save(output_file)

    print(f"Saved cleaned workbook: {output_file}")
    for sheet_name, status, deleted in summary:
        print(f"{sheet_name}: {status}, deleted_rows={deleted}")

    return output_file